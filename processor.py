"""
差分検出・DB保存モジュール（統合DB版）

シート構成:
  物件DB        ... 全条件横断のアクティブ/取消候補物件を1シートで管理
  成約・取消     ... 取消確定物件のアーカイブ
  変更ログ       ... 全変更履歴
"""

import json
import logging
import re
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

logger = logging.getLogger(__name__)

SHEET_DB      = "物件DB"
SHEET_REMOVED = "成約・取消"
SHEET_LOG     = "変更ログ"
ID_COL        = "物件番号"
PRICE_COL     = "価格"

STATUS_ACTIVE    = "アクティブ"
STATUS_CANDIDATE = "取消候補"

COLUMNS = [
    "物件番号", "物件種別", "取引状況", "取引態様",
    "所在地", "建物名", "所在階", "間取り",
    "専有面積", "建物面積", "土地面積",
    "価格", "㎡単価", "坪単価", "管理費",
    "用途地域", "建ぺい率", "容積率", "接道状況", "接道１",
    "沿線駅", "交通", "築年月",
    "会社名", "電話番号",
    "登録日", "図面", "検出条件", "状態", "取消候補日", "未検出回数",
    "グループID", "初回取得日", "最終確認日",
]
REMOVED_COLUMNS = COLUMNS + ["成約・取消日"]
LOG_COLUMNS     = ["日時", "検索条件名", "変更", "物件番号", "所在地", "価格", "旧価格"]


# ----------------------------------------------------------------
# 読み込み
# ----------------------------------------------------------------

def load_db(db_path: str) -> pd.DataFrame:
    """物件DBシートを読み込む。なければ空のDFを返す。"""
    if not Path(db_path).exists():
        return pd.DataFrame(columns=COLUMNS)
    try:
        xl = pd.ExcelFile(db_path)
        if SHEET_DB not in xl.sheet_names:
            return pd.DataFrame(columns=COLUMNS)
        df = pd.read_excel(xl, sheet_name=SHEET_DB, dtype=str).fillna("")
        # 旧形式DBに新カラムが無い場合は追加
        for col in COLUMNS:
            if col not in df.columns:
                df[col] = ""
        return df[COLUMNS]
    except Exception as e:
        logger.error(f"DB読込エラー: {e}")
        return pd.DataFrame(columns=COLUMNS)


def load_archive(db_path: str) -> pd.DataFrame:
    """成約・取消シートを読み込む。"""
    if not Path(db_path).exists():
        return pd.DataFrame(columns=REMOVED_COLUMNS)
    try:
        xl = pd.ExcelFile(db_path)
        if SHEET_REMOVED not in xl.sheet_names:
            return pd.DataFrame(columns=REMOVED_COLUMNS)
        df = pd.read_excel(xl, sheet_name=SHEET_REMOVED, dtype=str).fillna("")
        for col in REMOVED_COLUMNS:
            if col not in df.columns:
                df[col] = ""
        return df[REMOVED_COLUMNS]
    except Exception as e:
        logger.error(f"アーカイブ読込エラー: {e}")
        return pd.DataFrame(columns=REMOVED_COLUMNS)


def load_log(db_path: str) -> pd.DataFrame:
    if not Path(db_path).exists():
        return pd.DataFrame(columns=LOG_COLUMNS)
    try:
        xl = pd.ExcelFile(db_path)
        if SHEET_LOG not in xl.sheet_names:
            return pd.DataFrame(columns=LOG_COLUMNS)
        return pd.read_excel(xl, sheet_name=SHEET_LOG, dtype=str).fillna("")
    except Exception:
        return pd.DataFrame(columns=LOG_COLUMNS)


# ----------------------------------------------------------------
# バッチマージ（1回の実行で取れた全条件分の物件をDBに反映）
# ----------------------------------------------------------------

def merge_batch(
    db_df: pd.DataFrame,
    scraped_by_condition: list[tuple[str, list[dict]]],
    today: str,
    now_str: str,
    archive_df: pd.DataFrame | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, dict, list[dict]]:
    """
    複数条件分のスクレイプ結果をDBにマージする。
    archive_df が渡された場合、identityがアーカイブに一致したら成約・取消から復活させる。
    Returns:
      - 更新後のdb_df
      - 更新後のarchive_df（復活で除去された行が抜けたもの）
      - diff: {new, price_changed, restored, zumen_added, found_ids}
      - new_log_rows: 変更ログに追加する行
    """
    diff = {"new": [], "price_changed": [], "restored": [], "zumen_added": [], "found_ids": set()}
    log_rows: list[dict] = []

    # 全条件の物件を集めて、検索条件をまたいで会社名+丁目+徒歩分で
    # グループID を付け直す（同じ現場が条件をまたいで分割されてもまとめる）
    from rules import _mark_same_site_groups
    all_scraped_flat: list[dict] = []
    for _, props in scraped_by_condition:
        for p in props:
            p["グループID"] = ""
            all_scraped_flat.append(p)
    _mark_same_site_groups(all_scraped_flat)

    # DBの並び順を保つためレコードはリストで持ち、pid→index のマップで参照する
    records: list[dict] = [] if db_df.empty else db_df.to_dict("records")
    pid_to_idx: dict[str, int] = {}
    identity_to_idx: dict[tuple, int] = {}
    for i, rec in enumerate(records):
        pid = str(rec.get(ID_COL, "")).strip()
        if pid:
            pid_to_idx[pid] = i
        ikey = _identity_key(rec)
        if ikey is not None:
            identity_to_idx[ikey] = i

    # アーカイブ（成約・取消シート）のidentityインデックス
    archive_records: list[dict] = [] if archive_df is None or archive_df.empty else archive_df.to_dict("records")
    archive_identity_to_idx: dict[tuple, int] = {}
    for i, rec in enumerate(archive_records):
        ikey = _identity_key(rec)
        if ikey is not None and ikey not in archive_identity_to_idx:
            archive_identity_to_idx[ikey] = i
    archive_remove_indices: set[int] = set()

    def _apply_existing_update(rec: dict, prop: dict, condition_name: str,
                               report_price_change: bool = True) -> None:
        """既存レコードに今回のスクレイプ結果を反映し、価格変更・復活・図面追加を検知する。
        report_price_change=False の場合、価格は更新するが「価格変更」としては報告しない
        （アーカイブからの復活時は古い価格との差をノイズとして扱わない）。"""
        old_status = rec.get("状態", "")
        old_price  = rec.get(PRICE_COL, "")
        new_price  = prop.get(PRICE_COL, "")
        old_zumen  = rec.get("図面", "")
        new_zumen  = prop.get("図面", "")

        rec["最終確認日"] = today
        rec["状態"]        = STATUS_ACTIVE
        rec["取消候補日"] = ""
        rec["未検出回数"] = ""  # 見つかったので連続未検出カウントをリセット
        rec["検出条件"]   = _merge_conditions(rec.get("検出条件", ""), condition_name)
        if new_zumen:
            rec["図面"] = new_zumen

        if old_price and new_price and _norm_price(old_price) != _norm_price(new_price):
            rec[PRICE_COL] = new_price
            if report_price_change:
                p = dict(prop)
                p["旧価格"] = old_price
                p["新価格"] = new_price
                diff["price_changed"].append(p)
                log_rows.append(_log_row(now_str, condition_name, "価格変更", prop, old_price))

        if old_status == STATUS_CANDIDATE:
            diff["restored"].append(prop)
            log_rows.append(_log_row(now_str, condition_name, "取消候補から復活", prop))

        if old_zumen == "なし" and new_zumen == "あり":
            diff["zumen_added"].append(prop)
            log_rows.append(_log_row(now_str, condition_name, "図面追加", prop))

    # 今回のスクレイプに登場する全物件番号を事前計算
    # （再登録判定で「古い物件番号がまだ生きてるか」を見るため）
    all_found_pids: set[str] = set()
    for _, props in scraped_by_condition:
        for p in props:
            pp = str(p.get(ID_COL, "")).strip()
            if pp:
                all_found_pids.add(pp)

    pid_renewed = 0
    for condition_name, props in scraped_by_condition:
        for prop in props:
            pid = str(prop.get(ID_COL, "")).strip()
            if not pid:
                continue
            diff["found_ids"].add(pid)

            if pid in pid_to_idx:
                # 物件番号で既存ヒット → 位置はそのまま更新
                _apply_existing_update(records[pid_to_idx[pid]], prop, condition_name)
                continue

            # 物件番号は新規だが、同一物件キーで既存にヒットしないか確認（再登録検知）
            # ただし、ヒットした既存レコードの物件番号が今回のスクレイプにまだ存在する場合は
            # 「両方とも生きている別物件」なので再登録ではない → 真の新規として扱う
            ikey = _identity_key(prop)
            treat_as_distinct_new = False
            if ikey is not None and ikey in identity_to_idx:
                idx = identity_to_idx[ikey]
                rec = records[idx]
                old_pid     = str(rec.get(ID_COL, "")).strip()

                if old_pid and old_pid in all_found_pids:
                    # 旧物件番号もまだREINSにある → 別物件。再登録もアーカイブ復活も行わず新規扱い
                    treat_as_distinct_new = True
                else:
                    old_company = rec.get("会社名", "")
                    new_company = prop.get("会社名", "")

                    # 物件番号と会社名を最新に更新（位置は固定）
                    rec[ID_COL] = pid
                    if new_company:
                        rec["会社名"] = new_company
                    _apply_existing_update(rec, prop, condition_name)

                    # インデックスのpid参照を更新（位置は維持）
                    if old_pid in pid_to_idx:
                        del pid_to_idx[old_pid]
                    pid_to_idx[pid] = idx

                    # 変更ログには「物件番号変更」を残す（旧価格カラムに旧pidを格納）
                    log_rows.append(_log_row(
                        now_str, condition_name, "物件番号変更", prop, old_pid
                    ))
                    if new_company and old_company and new_company != old_company:
                        log_rows.append(_log_row(
                            now_str, condition_name, "会社名変更", prop, old_company
                        ))
                    pid_renewed += 1
                    continue

            # アーカイブ（成約・取消シート）に同一物件がないか確認
            # （別物件と判定済みの場合はアーカイブ復活もしない）
            if not treat_as_distinct_new and ikey is not None and ikey in archive_identity_to_idx:
                arch_idx = archive_identity_to_idx[ikey]
                arch_rec = archive_records[arch_idx]
                old_pid     = str(arch_rec.get(ID_COL, "")).strip()
                old_company = arch_rec.get("会社名", "")
                new_company = prop.get("会社名", "")

                # アーカイブ行をアクティブに戻す
                restored_rec = {col: arch_rec.get(col, "") for col in COLUMNS}
                restored_rec[ID_COL] = pid
                if new_company:
                    restored_rec["会社名"] = new_company
                restored_rec["状態"]        = STATUS_ACTIVE
                restored_rec["取消候補日"] = ""

                # アーカイブ復活時は価格変更として報告しない（古い価格との差はノイズ）
                _apply_existing_update(restored_rec, prop, condition_name, report_price_change=False)

                records.append(restored_rec)
                new_idx = len(records) - 1
                pid_to_idx[pid] = new_idx
                identity_to_idx[ikey] = new_idx
                archive_remove_indices.add(arch_idx)

                diff["restored"].append(prop)
                log_rows.append(_log_row(
                    now_str, condition_name, "アーカイブから復活", prop, old_pid
                ))
                if new_company and old_company and new_company != old_company:
                    log_rows.append(_log_row(
                        now_str, condition_name, "会社名変更", prop, old_company
                    ))
                continue

            # 真の新規物件
            new_rec = {col: prop.get(col, "") for col in COLUMNS}
            new_rec["検出条件"]   = condition_name
            new_rec["状態"]        = STATUS_ACTIVE
            new_rec["取消候補日"] = ""
            new_rec["初回取得日"] = today
            new_rec["最終確認日"] = today
            records.append(new_rec)
            new_idx = len(records) - 1
            pid_to_idx[pid] = new_idx
            if ikey is not None:
                identity_to_idx[ikey] = new_idx
            diff["new"].append(prop)
            log_rows.append(_log_row(now_str, condition_name, "新規登録", prop))

    # アーカイブから復活した行を除去
    if archive_remove_indices:
        archive_records = [r for i, r in enumerate(archive_records) if i not in archive_remove_indices]
    new_archive_df = (
        pd.DataFrame(archive_records, columns=REMOVED_COLUMNS)
        if archive_records else pd.DataFrame(columns=REMOVED_COLUMNS)
    )

    new_df = pd.DataFrame(records, columns=COLUMNS)
    logger.info(
        f"マージ完了 → 新規:{len(diff['new'])} "
        f"価格変更:{len(diff['price_changed'])} 復活:{len(diff['restored'])} "
        f"物件番号変更:{pid_renewed} アーカイブ復活:{len(archive_remove_indices)}"
    )
    return new_df, new_archive_df, diff, log_rows


# ----------------------------------------------------------------
# クリーンアップ: 既存DBに対する一括メンテナンス
# ----------------------------------------------------------------

def cleanup_db(db_path: str) -> dict:
    """
    既存DBの一括メンテナンス。
    1. グループID をグローバルに再計算
    2. identity が重複しているアクティブ行を統合（最古を残し、残りはアーカイブへ）

    バックアップは monitor.run_cleanup 側で作る前提。

    Returns: 統計情報 dict
    """
    db_df      = load_db(db_path)
    archive_df = load_archive(db_path)
    if db_df.empty:
        return {"active_before": 0, "active_after": 0, "merged": 0, "regrouped_count": 0}

    today   = datetime.now().strftime("%Y-%m-%d")
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    records = db_df.to_dict("records")
    active_before = sum(1 for r in records if r.get("状態", "") == STATUS_ACTIVE)

    # ── 重複統合 ──
    # identity が同じアクティブ行が複数あったら、最古（初回取得日が一番古い）を残す
    from collections import defaultdict
    ident_groups: dict[tuple, list[int]] = defaultdict(list)
    for i, rec in enumerate(records):
        if rec.get("状態", "") != STATUS_ACTIVE:
            continue
        ikey = _identity_key(rec)
        if ikey is not None:
            ident_groups[ikey].append(i)

    merged_count = 0
    archive_records = archive_df.to_dict("records") if not archive_df.empty else []
    log_rows: list[dict] = []
    to_archive_indices: set[int] = set()

    for ikey, idxs in ident_groups.items():
        if len(idxs) < 2:
            continue
        # 初回取得日昇順でソート → 最初のものをkeeperにする
        idxs_sorted = sorted(idxs, key=lambda i: str(records[i].get("初回取得日", "9999-99-99")))
        keeper = idxs_sorted[0]
        # keeper に他のレコードの検出条件をマージする
        for dup in idxs_sorted[1:]:
            r = records[dup]
            records[keeper]["検出条件"] = _merge_conditions(
                records[keeper].get("検出条件", ""),
                r.get("検出条件", "")
            )
            # 重複統合分はアーカイブに残さず削除する（変更ログにだけ記録）
            log_rows.append(_log_row(
                now_str, r.get("検出条件", ""), "重複統合(削除)", r, str(records[keeper].get(ID_COL, ""))
            ))
            to_archive_indices.add(dup)
            merged_count += 1

    if to_archive_indices:
        records = [r for i, r in enumerate(records) if i not in to_archive_indices]

    # ── グループID をグローバル再計算 ──
    from rules import _mark_same_site_groups
    active_only = [r for r in records if r.get("状態", "") == STATUS_ACTIVE]
    for r in active_only:
        r["グループID"] = ""
    _mark_same_site_groups(active_only)
    # active_only の dict 参照は records と共有してるので records 側も更新済み

    regrouped_count = sum(1 for r in active_only if r.get("グループID"))
    active_after = len(active_only)

    # ── アーカイブのゴミ掃除 ──
    # 成約・取消シートにあるが、同じ物件番号 or 同一identity がアクティブDBに存在する行は
    # 「実は成約してない（再登録で復活済み）」ので削除する。
    active_pids = {str(r.get(ID_COL, "")).strip() for r in records if str(r.get(ID_COL, "")).strip()}
    active_idents = set()
    for r in records:
        k = _identity_key(r)
        if k is not None:
            active_idents.add(k)

    kept_archive = []
    archive_orphans_removed = 0
    for r in archive_records:
        pid = str(r.get(ID_COL, "")).strip()
        ikey = _identity_key(r)
        if (pid and pid in active_pids) or (ikey is not None and ikey in active_idents):
            # アクティブに同一物件が存在 → 成約済みは誤りなので削除
            archive_orphans_removed += 1
            log_rows.append(_log_row(now_str, r.get("検出条件", ""), "誤成約取消を削除", r))
            continue
        kept_archive.append(r)
    archive_records = kept_archive

    new_db_df      = pd.DataFrame(records, columns=COLUMNS)
    new_archive_df = (
        pd.DataFrame(archive_records, columns=REMOVED_COLUMNS)
        if archive_records else pd.DataFrame(columns=REMOVED_COLUMNS)
    )
    save_db(db_path, new_db_df, new_archive_df, log_rows)

    return {
        "active_before":   active_before,
        "active_after":    active_after,
        "merged":          merged_count,
        "regrouped_count": regrouped_count,
        "archive_orphans_removed": archive_orphans_removed,
    }


# ----------------------------------------------------------------
# 週次: 取消候補マーキング
# ----------------------------------------------------------------

def mark_removal_candidates(
    db_df: pd.DataFrame, found_ids: set[str], today: str, now_str: str,
) -> tuple[pd.DataFrame, list[dict], list[dict]]:
    """
    今回の週次実行で見つからなかった物件を処理する。
    - アクティブで見つからなかった → 取消候補にして 未検出回数=1
    - 既に取消候補で今回も見つからなかった → 未検出回数 +1
    （見つかったものは merge_batch 側で 未検出回数 リセット済み）
    Returns:
      - 更新後のdb_df
      - candidates: 今回新たに取消候補になった物件dictのリスト
      - log_rows: 変更ログ行
    """
    candidates: list[dict] = []
    log_rows: list[dict]   = []
    if db_df.empty:
        return db_df, candidates, log_rows

    records = db_df.to_dict("records")
    incremented = 0
    for rec in records:
        pid    = str(rec.get(ID_COL, "")).strip()
        status = rec.get("状態", "")
        if not pid or pid in found_ids:
            continue
        if status == STATUS_ACTIVE:
            # 新たに見つからなくなった → 取消候補1回目
            rec["状態"]        = STATUS_CANDIDATE
            rec["取消候補日"] = today
            rec["未検出回数"] = "1"
            candidates.append(dict(rec))
            log_rows.append(_log_row(now_str, rec.get("検出条件", ""), "取消候補(1回目)", rec))
        elif status == STATUS_CANDIDATE:
            # 既に取消候補で今回も見つからない → カウントアップ
            try:
                cnt = int(str(rec.get("未検出回数", "0") or "0"))
            except ValueError:
                cnt = 0
            cnt += 1
            rec["未検出回数"] = str(cnt)
            incremented += 1
            log_rows.append(_log_row(now_str, rec.get("検出条件", ""), f"取消候補({cnt}回目)", rec))

    new_df = pd.DataFrame(records, columns=COLUMNS)
    logger.info(f"取消候補マーキング: 新規{len(candidates)}件 / 継続未検出{incremented}件")
    return new_df, candidates, log_rows


# ----------------------------------------------------------------
# 猶予期間切れ: 取消候補を成約・取消シートへ移動
# ----------------------------------------------------------------

def process_grace_period(
    db_df: pd.DataFrame,
    archive_df: pd.DataFrame,
    today: str,
    now_str: str,
    confirm_misses: int = 2,
) -> tuple[pd.DataFrame, pd.DataFrame, list[dict], list[dict]]:
    """
    取消候補のうち、週次実行で confirm_misses 回連続見つからなかった物件を
    成約・取消シートへ移動する（=成約確定）。
    Returns:
      - 更新後のdb_df（移動した物件は除去）
      - 更新後のarchive_df（移動した物件を追加）
      - confirmed: 成約・取消が確定した物件dictリスト
      - log_rows: 変更ログ行
    """
    confirmed: list[dict] = []
    log_rows: list[dict]   = []
    if db_df.empty:
        return db_df, archive_df, confirmed, log_rows

    keep_records: list[dict] = []
    archive_records = archive_df.to_dict("records") if not archive_df.empty else []

    for rec in db_df.to_dict("records"):
        if rec.get("状態", "") != STATUS_CANDIDATE:
            keep_records.append(rec)
            continue
        try:
            misses = int(str(rec.get("未検出回数", "0") or "0"))
        except ValueError:
            misses = 0
        if misses >= confirm_misses:
            arch = {col: rec.get(col, "") for col in COLUMNS}
            arch["成約・取消日"] = today
            archive_records.append(arch)
            confirmed.append(dict(rec))
            log_rows.append(_log_row(now_str, rec.get("検出条件", ""), "成約・取消確定", rec))
        else:
            keep_records.append(rec)

    new_db_df      = pd.DataFrame(keep_records,    columns=COLUMNS)
    new_archive_df = pd.DataFrame(archive_records, columns=REMOVED_COLUMNS)
    if confirmed:
        logger.info(f"成約・取消確定: {len(confirmed)}件（猶予{grace_days}日経過）")
    return new_db_df, new_archive_df, confirmed, log_rows


# ----------------------------------------------------------------
# 取消候補 → アクティブ 復元（戻す.bat 用）
# ----------------------------------------------------------------

def restore_candidate(db_path: str, prop_id: str) -> bool:
    """単一物件の取消候補→アクティブ。"""
    ok, _ = restore_candidates(db_path, [prop_id])
    return ok > 0


def restore_candidates(db_path: str, prop_ids: list[str]) -> tuple[int, list[str]]:
    """
    複数物件をまとめて取消候補→アクティブに戻す。
    Returns: (成功件数, 見つからなかった物件番号リスト)
    """
    if not prop_ids:
        return 0, []
    db_df = load_db(db_path)
    if db_df.empty:
        return 0, list(prop_ids)

    targets = {str(p).strip() for p in prop_ids if str(p).strip()}
    series  = db_df[ID_COL].astype(str).str.strip()
    mask    = series.isin(targets)
    if not mask.any():
        return 0, list(targets)

    db_df.loc[mask, "状態"]       = STATUS_ACTIVE
    db_df.loc[mask, "取消候補日"] = ""
    save_db(db_path, db_df, load_archive(db_path), [])

    found = set(series[mask])
    not_found = sorted(targets - found)
    return int(mask.sum()), not_found


def confirm_removal(db_path: str, prop_id: str) -> bool:
    """単一物件を物件DBから外して成約・取消シートへ。"""
    ok, _ = confirm_removals(db_path, [prop_id])
    return ok > 0


def confirm_removals(db_path: str, prop_ids: list[str]) -> tuple[int, list[str]]:
    """
    複数物件をまとめて成約・取消シートへ移動。
    Returns: (成功件数, 見つからなかった物件番号リスト)
    """
    if not prop_ids:
        return 0, []
    db_df      = load_db(db_path)
    archive_df = load_archive(db_path)
    if db_df.empty:
        return 0, list(prop_ids)

    targets = {str(p).strip() for p in prop_ids if str(p).strip()}
    series  = db_df[ID_COL].astype(str).str.strip()
    mask    = series.isin(targets)
    if not mask.any():
        return 0, list(targets)

    today   = datetime.now().strftime("%Y-%m-%d")
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    moving_rows = db_df[mask].to_dict("records")
    archive_records = archive_df.to_dict("records") if not archive_df.empty else []
    log_rows: list[dict] = []
    for rec in moving_rows:
        arch = {col: rec.get(col, "") for col in COLUMNS}
        arch["成約・取消日"] = today
        archive_records.append(arch)
        log_rows.append(_log_row(now_str, rec.get("検出条件", ""), "成約・取消（手動）", rec))

    new_db_df  = db_df[~mask].reset_index(drop=True)
    new_arch_df = pd.DataFrame(archive_records, columns=REMOVED_COLUMNS)
    save_db(db_path, new_db_df, new_arch_df, log_rows)

    found = set(series[mask])
    not_found = sorted(targets - found)
    return int(mask.sum()), not_found


# ----------------------------------------------------------------
# 保存
# ----------------------------------------------------------------

def save_db(
    db_path: str,
    db_df: pd.DataFrame,
    archive_df: pd.DataFrame,
    new_log_rows: list[dict],
) -> None:
    """物件DB・成約取消・変更ログをまとめて保存する。"""
    log_df = load_log(db_path)
    log_records = log_df.to_dict("records") if not log_df.empty else []
    log_records.extend(new_log_rows)
    log_full = (
        pd.DataFrame(log_records, columns=LOG_COLUMNS)
        if log_records
        else pd.DataFrame(columns=LOG_COLUMNS)
    )

    with pd.ExcelWriter(db_path, engine="openpyxl") as writer:
        db_df.to_excel(writer,      sheet_name=SHEET_DB,      index=False)
        archive_df.to_excel(writer, sheet_name=SHEET_REMOVED, index=False)
        log_full.to_excel(writer,   sheet_name=SHEET_LOG,     index=False)

    _apply_styles(db_path)
    logger.info(f"DB保存完了: {db_path} アクティブ{(db_df['状態']==STATUS_ACTIVE).sum() if not db_df.empty else 0}件 取消候補{(db_df['状態']==STATUS_CANDIDATE).sum() if not db_df.empty else 0}件")


# ----------------------------------------------------------------
# 内部ヘルパー
# ----------------------------------------------------------------

def _merge_conditions(existing: str, new_cond: str) -> str:
    parts = [s.strip() for s in str(existing).split(",") if s.strip()]
    if new_cond and new_cond not in parts:
        parts.append(new_cond)
    return ", ".join(parts)


def _norm_num(s) -> str:
    """数値文字列を正規化（72.40 → 72.4、空白・単位除去）。"""
    s = re.sub(r"[\s,、　円万㎡m2平米]", "", str(s or "")).strip()
    if not s:
        return ""
    try:
        return f"{float(s):g}"
    except ValueError:
        return s


# カタカナ濁点・半濁点の除去テーブル（ザ→サ、グ→ク、パ→ハなど）
_KATAKANA_DAKUTEN_TABLE = str.maketrans(
    "ガギグゲゴザジズゼゾダヂヅデドバビブベボパピプペポヴ",
    "カキクケコサシスセソタチツテトハヒフヘホハヒフヘホウ",
)


def _norm_text(s) -> str:
    """文字列を識別キー用に正規化。
    - NFKC正規化で 全角英数→半角、互換文字を統一
    - 空白・記号類を除去
    - カタカナの濁点・半濁点を除去（表記ゆれ吸収：ザ↔サなど）
    """
    s = unicodedata.normalize("NFKC", str(s or ""))
    s = re.sub(r"[\s\-‐‑‒–—―ーｰ・.,()\[\]{}!?#&/＋+]", "", s)
    s = s.translate(_KATAKANA_DAKUTEN_TABLE)
    return s


def _norm_floor(s) -> str:
    """所在階を数字（および地下のB）だけに整形。"6階" → "6", "B1階" → "B1"。"""
    s = unicodedata.normalize("NFKC", str(s or "")).upper()
    m = re.search(r"(B?\d+)", s)
    return m.group(1) if m else ""


def _norm_type(s) -> str:
    """物件種別を正規化。"中古マンション"・"新築マンション" → "マンション" など。"""
    t = _norm_text(s)
    t = re.sub(r"^(中古|新築|売|新)", "", t)
    return t


def _identity_key(rec: dict) -> tuple | None:
    """
    物件を一意識別するキー（物件番号・価格・会社名を除く）。
    所在地が無いと識別できないので None を返す。
    会社名は含めないので、別会社が同じ物件を再登録した場合も同一物件と判定する。
    物件種別は中古/新築の差を無視。所在地・建物名は空白を除いて表記ゆれを吸収。
    """
    addr = _norm_text(rec.get("所在地"))
    if not addr:
        return None
    return (
        _norm_type(rec.get("物件種別")),
        addr,
        _norm_text(rec.get("建物名")),
        _norm_floor(rec.get("所在階")),
        _norm_text(rec.get("間取り")),
        _norm_num(rec.get("専有面積")),
        _norm_num(rec.get("建物面積")),
        _norm_num(rec.get("土地面積")),
        _norm_text(rec.get("築年月")),
    )


def _log_row(now: str, cond: str, change: str, prop: dict, old_price: str = "") -> dict:
    return {
        "日時":      now,
        "検索条件名": cond,
        "変更":      change,
        "物件番号":   prop.get(ID_COL, ""),
        "所在地":     prop.get("所在地", ""),
        "価格":       prop.get(PRICE_COL, ""),
        "旧価格":     old_price,
    }


def _norm_price(s: str) -> str:
    return re.sub(r"[\s,、　円万]", "", str(s))


def _apply_styles(db_path: str) -> None:
    try:
        wb = openpyxl.load_workbook(db_path)
        if SHEET_DB not in wb.sheetnames:
            wb.save(db_path)
            return
        ws = wb[SHEET_DB]

        fill_candidate = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        fill_group     = PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid")

        for cell in ws[1]:
            cell.font      = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # 列インデックス
        headers   = [c.value for c in ws[1]]
        status_i  = headers.index("状態")     + 1 if "状態"     in headers else None
        group_i   = headers.index("グループID") + 1 if "グループID" in headers else None

        for row in ws.iter_rows(min_row=2):
            status = row[status_i - 1].value if status_i else None
            grp    = row[group_i - 1].value if group_i else None
            if status == STATUS_CANDIDATE:
                for c in row:
                    c.fill = fill_candidate
            elif grp:
                for c in row:
                    c.fill = fill_group

        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(w + 2, 45)

        wb.save(db_path)
    except Exception as e:
        logger.warning(f"スタイル適用エラー（無視）: {e}")


# ----------------------------------------------------------------
# 実行状態（最終実行日など）の永続化
# ----------------------------------------------------------------

def load_state(state_path: str) -> dict:
    """状態ファイルを読み込む。なければ空dictを返す。"""
    p = Path(state_path)
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:
        logger.warning(f"state読込失敗（空として扱います）: {e}")
        return {}


def save_state(state_path: str, state: dict) -> None:
    """状態ファイルを書き込む。"""
    try:
        Path(state_path).write_text(
            json.dumps(state, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    except Exception as e:
        logger.warning(f"state保存失敗: {e}")
